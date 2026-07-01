from datetime import datetime, timedelta
import csv
import io
import logging
import secrets
from urllib.parse import urlencode

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response

from backend.src.config.frontend import templates
from backend.src.config.settings import settings
from backend.src.middleware.auth import get_current_user
from backend.src.middleware.auth import pwd_context
from backend.src.middleware.csrf import create_csrf_token, set_csrf_cookie, validate_csrf_token
from backend.src.roles import ROLE_LABELS, ROLES, SUPERVISOR, USER, can_manage_users, can_review_art
from backend.src.services.art_service import (
    actualizar_revision_art,
    cargar_registros,
    cargar_registros_por_supervisor,
    cargar_nombres_asignaciones,
    cargar_resumenes_asignaciones,
    obtener_registro,
)
from backend.src.services.content_filter import (
    PROHIBITED_LANGUAGE_MESSAGE,
    find_prohibited_terms,
    validate_clean_fields,
)
from backend.src.services.usuario_service import (
    actualizar_estado_cuenta,
    actualizar_rol,
    cargar_usuarios,
    email_existe,
    guardar_activation_token,
    guardar_usuario,
    obtener_usuario,
    rut_existe,
    username_existe,
    obtener_usernames_por_ids,
)
from backend.src.services.validation_service import (
    normalizar_rut,
    normalizar_telefono_chile,
    validar_rut_chileno,
    validar_telefono_chile,
)
from backend.src.services.pdf_service import generar_art_pdf
from backend.src.services.notification_service import add_notification
from backend.src.services.realtime_service import realtime_manager
from backend.src.services.email_service import render_email_template, render_text_email, send_activation_email, send_email
from backend.src.config.database import _connect

router = APIRouter()
logger = logging.getLogger("dart.admin")
IMPORT_COLUMNS = ["nombre", "apellido", "rut", "email", "rol", "telefono", "cargo", "area"]
IMPORT_RESULT_COLUMNS = [
    "nombre",
    "apellido",
    "rut",
    "email",
    "rol",
    "telefono",
    "cargo",
    "area",
    "estado",
    "detalle",
    "email_status",
]


def _cuenta_activa(user: dict | None) -> bool:
    if not user:
        return False
    return user.get("estado_cuenta") == "activo" or user.get("estado") == "activo" or user.get("is_active") is True


def _email_corporativo_valido(email: str) -> bool:
    email = email.strip().lower()
    if "@" not in email:
        return False
    return email.rsplit("@", 1)[-1] in settings.allowed_email_domains


def _username_desde_email(email: str) -> str:
    base = "".join(ch for ch in email.split("@", 1)[0].lower() if ch.isalnum() or ch in "._-")[:30] or "usuario"
    username = base
    while username_existe(username):
        username = f"{base[:25]}-{secrets.token_hex(2)}"
    return username


def _absolute_url(request: Request, path: str) -> str:
    base = settings.public_base_url or "https://www.dart-mineria.lat"
    return f"{base}{path}"


def _crear_activation_link(request: Request, username: str) -> str:
    target = obtener_usuario(username)
    if _cuenta_activa(target):
        logger.warning("ACTIVATION_LINK_BLOCKED_ALREADY_ACTIVE username=%s", username)
        raise HTTPException(status_code=400, detail="Cuenta ya está activada")
    token = secrets.token_urlsafe(32)
    if not guardar_activation_token(username, token, datetime.now() + timedelta(hours=48)):
        logger.warning("USER_ALREADY_ACTIVE username=%s", username)
        raise HTTPException(status_code=400, detail="Cuenta ya está activada")
    link = _absolute_url(request, f"/activar-cuenta/{token}")
    logger.info("activation_token_created username=%s", username)
    return link


def _redirect_admin_usuarios(**params: str) -> RedirectResponse:
    query = urlencode({key: value for key, value in params.items() if value})
    url = "/admin/usuarios"
    if query:
        url = f"{url}?{query}"
    return RedirectResponse(url, status_code=303)


def _enviar_activacion_usuario(email: str, link: str):
    return send_activation_email(email, link)


def _enviar_email_prueba(email: str, link: str) -> bool:
    subject = "Prueba de correo D.A.R.T"
    intro = "Este correo confirma que el envío de D.A.R.T mediante Resend está configurado correctamente."
    html_body = render_email_template(
        title="Prueba de correo D.A.R.T",
        intro=intro,
        action_text="Abrir D.A.R.T",
        action_url=link,
        footer_note="Correo enviado desde la herramienta de prueba de administración.",
    )
    text_body = render_text_email(
        "Prueba de correo D.A.R.T",
        intro,
        "Abrir D.A.R.T",
        link,
        "Correo enviado desde la herramienta de prueba de administración.",
    )
    return send_email(email, subject, html_body, text_body)


def _parse_import_file(filename: str, content: bytes) -> list[dict[str, str]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        return [dict(row) for row in csv.DictReader(io.StringIO(text))]
    if lower_name.endswith((".xlsx", ".xlsm", ".xls")):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=400, detail="Para importar Excel instala openpyxl.") from exc
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip().lower() for value in rows[0]]
        data = []
        for values in rows[1:]:
            data.append({headers[idx]: str(value or "").strip() for idx, value in enumerate(values) if idx < len(headers)})
        return data
    raise HTTPException(status_code=400, detail="El archivo debe ser CSV o Excel.")


def _admin_context(request: Request, user: dict, **extra):
    return {
        "request": request,
        "user": user,
        "usuarios": cargar_usuarios(),
        "roles": ROLE_LABELS,
        "title": "Gestionar usuarios",
        "allowed_domains": sorted(settings.allowed_email_domains),
        "status_message": request.query_params.get("message"),
        "status_type": request.query_params.get("type", "info"),
        **extra,
    }


def _render_admin_usuarios(request: Request, user: dict, **extra):
    csrf_token = create_csrf_token()
    response = templates.TemplateResponse(
        request,
        "admin_users.html",
        _admin_context(request, user, csrf_token=csrf_token, **extra),
    )
    set_csrf_cookie(response, csrf_token)
    return response


def _validar_datos_usuario(
    nombre: str,
    apellido: str,
    rut: str,
    email: str,
    rol: str,
    telefono: str,
    cargo: str = "",
    area: str = "",
    validar_duplicados: bool = True,
) -> tuple[dict[str, str], dict[str, str]]:
    errors: dict[str, str] = {}
    normalized = {
        "nombre": nombre.strip(),
        "apellido": apellido.strip(),
        "rut": normalizar_rut(rut),
        "email": email.strip().lower(),
        "rol": rol.strip().lower(),
        "telefono": normalizar_telefono_chile(telefono),
    }
    if not normalized["nombre"]:
        errors["nombre"] = "El nombre es obligatorio."
    if not normalized["rut"]:
        errors["rut"] = "El RUT es obligatorio."
    elif not validar_rut_chileno(normalized["rut"]):
        errors["rut"] = "El RUT ingresado no es válido."
    elif validar_duplicados and rut_existe(normalized["rut"]):
        errors["rut"] = "El RUT ya está registrado."
    if not normalized["email"]:
        errors["email"] = "El email es obligatorio."
    elif not _email_corporativo_valido(normalized["email"]):
        errors["email"] = "Solo se aceptan correos @dart-mineria.lat."
    elif validar_duplicados and email_existe(normalized["email"]):
        errors["email"] = "Ya existe un usuario con ese correo."
    if normalized["rol"] not in ROLES:
        errors["rol"] = "Rol inválido. Usa admin, supervisor o trabajador."
    if telefono.strip() and not validar_telefono_chile(telefono):
        errors["telefono"] = "El teléfono debe tener formato chileno válido: +56 9 XXXX XXXX."
    for field_name, value in {"nombre": nombre, "apellido": apellido, "cargo": cargo, "area": area}.items():
        if find_prohibited_terms(value):
            errors[field_name] = PROHIBITED_LANGUAGE_MESSAGE
    return errors, normalized


def _es_supervisor_asignado(registro: dict, user: dict) -> bool:
    if user.get("rol") != SUPERVISOR:
        return False
    asignado = (registro.get("supervisor_asignado") or "").strip().lower()
    if asignado:
        return asignado == user.get("username", "").strip().lower()
    supervisor_texto = (registro.get("supervisor") or "").strip().lower()
    opciones = {
        user.get("username", "").strip().lower(),
        user.get("nombre", "").strip().lower(),
        user.get("email", "").strip().lower(),
    }
    return supervisor_texto in opciones


@router.get("/admin/art", response_class=HTMLResponse)
@router.get("/supervisor/art", response_class=HTMLResponse)
def admin_list_art(request: Request, user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_review_art(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    registros = cargar_registros(limite=100, con_asignaciones=False) if user.get("rol") == "admin" else cargar_registros_por_supervisor(user["username"], limite=100, con_asignaciones=False)
    resumenes = cargar_resumenes_asignaciones([registro["id"] for registro in registros])
    nombres_por_art = cargar_nombres_asignaciones([registro["id"] for registro in registros])
    for registro in registros:
        registro["progreso_asignaciones"] = resumenes.get(registro["id"], {})
        registro["trabajadores_nombres"] = nombres_por_art.get(registro["id"], [])
    csrf_token = create_csrf_token()
    response = templates.TemplateResponse(
        request,
        "admin_art_list.html",
        {
            "request": request,
            "user": user,
            "registros": registros,
            "title": "Revisar ARTs",
            "csrf_token": csrf_token,
            "can_manage_users": can_manage_users(user.get("rol", "")),
        },
    )
    set_csrf_cookie(response, csrf_token)
    return response


@router.post("/admin/art/{id_art}/estado")
@router.post("/supervisor/art/{id_art}/estado")
async def admin_change_estado(
    request: Request,
    id_art: str,
    estado: str = Form(...),
    comentario_supervisor: str = Form(""),
    csrf_token: str = Form(...),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_review_art(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    validate_csrf_token(request, csrf_token)
    registro = obtener_registro(id_art)
    if not registro:
        return RedirectResponse("/supervisor/art", status_code=303)
    if user.get("rol") != "admin" and not _es_supervisor_asignado(registro, user):
        return RedirectResponse("/dashboard", status_code=303)
    # Prevent changing estado once it's finalized
    if registro.get("estado") in {"aprobada", "rechazada", "completada"}:
        return RedirectResponse(f"/art/{id_art}", status_code=303)
    if estado not in {"pendiente", "aprobada", "rechazada"}:
        raise HTTPException(status_code=400, detail="Estado de ART inválido")
    if estado == "rechazada" and not comentario_supervisor.strip():
        return JSONResponse(
            {
                "error": "validation_error",
                "field": "comentario",
                "message": "Comentario obligatorio para rechazar ART",
            },
            status_code=400,
        )
    if estado == "aprobada" and not comentario_supervisor.strip():
        raise HTTPException(status_code=400, detail="Debes ingresar un comentario de revisión")
    validate_clean_fields({"comentario_supervisor": comentario_supervisor}, user.get("username", ""))
        
    actualizar_revision_art(
        id_art,
        estado,
        comentario_supervisor.strip(),
        user["username"],
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    try:
        target_usernames = {registro.get("creado_por", "")}
        worker_ids = [item.get("trabajador_id") for item in registro.get("asignaciones", [])]
        target_usernames.update(obtener_usernames_por_ids(worker_ids))
        for target_username in target_usernames - {""}:
            notification = add_notification(
                target_username,
                f"ART {id_art} {estado}",
                comentario_supervisor.strip() or "Se actualizó el estado de la ART",
                f"/art/{id_art}",
                "ART_STATUS_CHANGE",
            )
            await realtime_manager.send_notification(notification)
    except Exception:
        logger.exception("art_status_notification_failed art_id=%s estado=%s", id_art, estado)
    # If resolved (approved/rejected), generate PDF backup into uploads
    if estado in {"aprobada", "rechazada"}:
        try:
            registro_actual = obtener_registro(id_art)
            pdf_bytes = generar_art_pdf(registro_actual)  # registro ya trae sus asignaciones
            upload_dir = request.app.state.upload_dir
            (upload_dir / f"art-{id_art}.pdf").write_bytes(pdf_bytes)
        except Exception:
            # don't block the flow if PDF generation fails
            pass
    return RedirectResponse(f"/art/{id_art}", status_code=303)


@router.get("/admin/usuarios", response_class=HTMLResponse)
def admin_list_usuarios(request: Request, user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_manage_users(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    return _render_admin_usuarios(request, user)


@router.get("/admin/usuarios/plantilla")
def admin_descargar_plantilla(user=Depends(get_current_user)):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_manage_users(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=IMPORT_COLUMNS)
    writer.writeheader()
    writer.writerow(
        {
            "nombre": "Juan",
            "apellido": "González",
            "rut": "12.345.678-5",
            "email": "juan.perez@dart-mineria.lat",
            "rol": "trabajador",
            "telefono": "+56 9 1234 5678",
            "cargo": "Operador",
            "area": "Prevención",
        }
    )
    writer.writerow(
        {
            "nombre": "María",
            "apellido": "Soto",
            "rut": "11.111.111-1",
            "email": "maria.soto@dart-mineria.lat",
            "rol": "supervisor",
            "telefono": "+56 9 8765 4321",
            "cargo": "Supervisora",
            "area": "Operaciones",
        }
    )
    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="plantilla-usuarios-dart.csv"'},
    )


@router.post("/admin/usuarios/{username}/rol")
def admin_update_rol(
    request: Request,
    username: str,
    rol: str = Form(...),
    csrf_token: str = Form(...),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_manage_users(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    validate_csrf_token(request, csrf_token)
    if rol not in ROLES:
        raise HTTPException(status_code=400, detail="Rol inválido")
    if username == user["username"] and rol != "admin":
        raise HTTPException(status_code=400, detail="No puedes quitarte tu propio rol admin")
    actualizar_rol(username, rol)
    try:
        from backend.src.services import logging_service as _log_svc
        _log_svc.log_event(_log_svc.USER_ROLE_CHANGED, username=user.get("username", ""), details={"target": username, "new_rol": rol})
    except Exception:
        pass
    return RedirectResponse("/admin/usuarios", status_code=303)


@router.post("/admin/usuarios/crear")
def admin_crear_usuario(
    request: Request,
    nombre: str = Form(...),
    apellido: str = Form(""),
    rut: str = Form(...),
    email: str = Form(...),
    rol: str = Form(USER),
    telefono: str = Form(""),
    cargo: str = Form(""),
    area: str = Form(""),
    csrf_token: str = Form(...),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_manage_users(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    validate_csrf_token(request, csrf_token)
    form_data = {
        "nombre": nombre,
        "apellido": apellido,
        "rut": rut,
        "email": email,
        "rol": rol,
        "telefono": telefono,
        "cargo": cargo,
        "area": area,
    }
    try:
        validate_clean_fields({"nombre": nombre, "apellido": apellido, "cargo": cargo, "area": area}, user.get("username", ""))
    except HTTPException:
        return _render_admin_usuarios(
            request,
            user,
            form_errors={"contenido": PROHIBITED_LANGUAGE_MESSAGE},
            form_data=form_data,
        )
    form_errors, normalized = _validar_datos_usuario(nombre, apellido, rut, email, rol, telefono, cargo, area)
    if form_errors:
        return _render_admin_usuarios(request, user, form_errors=form_errors, form_data=form_data)
    username = _username_desde_email(normalized["email"])
    guardar_usuario(
        username=username,
        password_hash=pwd_context.hash(secrets.token_urlsafe(24)),
        rol=normalized["rol"],
        nombre=normalized["nombre"],
        apellido=normalized["apellido"],
        email=normalized["email"],
        rut=normalized["rut"],
        telefono=normalized["telefono"],
        cargo=cargo.strip(),
        area=area.strip(),
        estado_cuenta="pendiente",
        debe_cambiar_password=True,
    )
    link = _crear_activation_link(request, username)
    result = _enviar_activacion_usuario(normalized["email"], link)
    if result.ok:
        return _redirect_admin_usuarios(message="Usuario creado. Se envió el enlace de activación al correo.", type="success")
    logger.error("EMAIL_SENT_FAIL context=activation username=%s email=%s error=%s", username, normalized["email"], result.error)
    return _redirect_admin_usuarios(
        message="Usuario creado, pero falló el envío del correo. Revisa configuración Resend/EMAIL_ENABLED y reenvía la activación.",
        type="warning",
    )


@router.post("/admin/usuarios/{username}/estado")
def admin_update_estado(
    request: Request,
    username: str,
    estado: str = Form(...),
    csrf_token: str = Form(...),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_manage_users(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    validate_csrf_token(request, csrf_token)
    if estado not in {"activo", "inactivo", "pendiente"}:
        raise HTTPException(status_code=400, detail="Estado inválido")
    if username == user["username"] and estado != "activo":
        raise HTTPException(status_code=400, detail="No puedes desactivar tu propia cuenta")
    actualizar_estado_cuenta(username, estado)
    return RedirectResponse("/admin/usuarios", status_code=303)


@router.post("/admin/usuarios/{username}/reenviar-activacion")
def admin_reenviar_activacion(
    request: Request,
    username: str,
    csrf_token: str = Form(...),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_manage_users(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    validate_csrf_token(request, csrf_token)
    target = obtener_usuario(username)
    if not target:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if _cuenta_activa(target):
        logger.warning("USER_ALREADY_ACTIVE username=%s", username)
        return _redirect_admin_usuarios(message="Cuenta ya está activada", type="info")
    link = _crear_activation_link(request, username)
    result = _enviar_activacion_usuario(target.get("email", ""), link)
    if result.ok:
        return _redirect_admin_usuarios(message="Se envió un nuevo enlace de activación.", type="success")
    logger.error("EMAIL_SENT_FAIL context=activation_resend username=%s email=%s error=%s", username, target.get("email", ""), result.error)
    return _redirect_admin_usuarios(message="No se pudo enviar el correo de activación. Revisa Resend/EMAIL_ENABLED.", type="warning")


@router.post("/admin/test-email", name="admin_test_email")
def admin_test_email(
    request: Request,
    csrf_token: str = Form(...),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_manage_users(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    validate_csrf_token(request, csrf_token)
    email = (user.get("email") or "").strip()
    if not email:
        return _redirect_admin_usuarios(message="Tu usuario admin no tiene email registrado.", type="warning")
    test_link = _absolute_url(request, "/dashboard")
    if _enviar_email_prueba(email, test_link):
        return _redirect_admin_usuarios(message="Correo de prueba enviado al admin actual.", type="success")
    return _redirect_admin_usuarios(message="No se pudo enviar el correo de prueba. Revisa Resend en Railway.", type="warning")


@router.post("/admin/usuarios/importar", response_class=HTMLResponse)
async def admin_importar_usuarios(
    request: Request,
    archivo: UploadFile = File(...),
    csrf_token: str = Form(...),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_manage_users(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    validate_csrf_token(request, csrf_token)
    try:
        rows = _parse_import_file(archivo.filename or "", await archivo.read())
    except (HTTPException, UnicodeDecodeError) as exc:
        detail = exc.detail if isinstance(exc, HTTPException) else "No se pudo leer el archivo. Usa CSV UTF-8 o Excel válido."
        return _render_admin_usuarios(request, user, import_error=detail)
    if not rows:
        return _render_admin_usuarios(request, user, import_error="El archivo no contiene usuarios para importar.")
    headers = {key.strip().lower() for row in rows for key in row.keys() if key}
    missing_columns = [column for column in ["nombre", "rut", "email", "rol"] if column not in headers]
    if missing_columns:
        return _render_admin_usuarios(
            request,
            user,
            import_error=f"Faltan columnas obligatorias: {', '.join(missing_columns)}.",
        )
    resultados = []
    creados = duplicados = errores = 0
    correos_enviados = correos_fallidos = 0
    ruts_en_archivo: set[str] = set()
    emails_en_archivo: set[str] = set()
    for index, row in enumerate(rows, start=2):
        nombre = (row.get("nombre") or "").strip()
        apellido = (row.get("apellido") or "").strip()
        rut_input = (row.get("rut") or "").strip()
        email = (row.get("email") or "").strip().lower()
        rol = (row.get("rol") or USER).strip().lower()
        telefono_input = (row.get("telefono") or "").strip()
        cargo = (row.get("cargo") or "").strip()
        area = (row.get("area") or "").strip()
        detalle = ""
        link = ""
        email_status = ""
        estado = "error"
        row_errors, normalized = _validar_datos_usuario(
            nombre,
            apellido,
            rut_input,
            email,
            rol,
            telefono_input,
            cargo,
            area,
            validar_duplicados=False,
        )
        try:
            validate_clean_fields(
                {"nombre": nombre, "apellido": apellido, "cargo": cargo, "area": area},
                user.get("username", ""),
            )
        except HTTPException:
            row_errors["contenido"] = PROHIBITED_LANGUAGE_MESSAGE
        rut_normalizado = normalized["rut"]
        email_normalizado = normalized["email"]
        telefono_normalizado = normalized["telefono"]
        if not row_errors and rut_normalizado in ruts_en_archivo:
            row_errors["rut"] = "RUT duplicado dentro del archivo."
        if not row_errors and email_normalizado in emails_en_archivo:
            row_errors["email"] = "Email duplicado dentro del archivo."
        if not row_errors and rut_existe(rut_normalizado):
            estado = "duplicado"
            detalle = "RUT ya registrado."
            duplicados += 1
        elif not row_errors and email_existe(email_normalizado):
            estado = "duplicado"
            detalle = "Email ya registrado."
            duplicados += 1
        if row_errors:
            detalle = " ".join(row_errors.values())
            errores += 1
        elif estado != "duplicado":
            username = _username_desde_email(email_normalizado)
            guardar_usuario(
                username=username,
                password_hash=pwd_context.hash(secrets.token_urlsafe(24)),
                rol=normalized["rol"],
                nombre=normalized["nombre"],
                apellido=normalized["apellido"],
                email=email_normalizado,
                rut=rut_normalizado,
                telefono=telefono_normalizado,
                cargo=cargo,
                area=area,
                estado_cuenta="pendiente",
                debe_cambiar_password=True,
            )
            link = _crear_activation_link(request, username)
            estado = "creado"
            detalle = "Usuario creado en estado pendiente."
            email_status = "pendiente"
            creados += 1
            result = _enviar_activacion_usuario(email_normalizado, link)
            link = ""
            if result.ok:
                correos_enviados += 1
                email_status = "email_sent"
                detalle = "Usuario creado. Correo de activación enviado."
            else:
                correos_fallidos += 1
                email_status = "email_failed"
                detalle = "Usuario creado, pero falló el envío del correo."
                logger.error("EMAIL_SENT_FAIL context=activation_import username=%s email=%s error=%s", username, email_normalizado, result.error)
        if not row_errors and rut_normalizado:
            ruts_en_archivo.add(rut_normalizado)
        if not row_errors and email_normalizado:
            emails_en_archivo.add(email_normalizado)
        resultados.append(
            {
                "fila": index,
                "nombre": normalized["nombre"] or nombre,
                "apellido": normalized["apellido"] or apellido,
                "rut": rut_normalizado,
                "email": email_normalizado,
                "rol": normalized["rol"] or rol,
                "telefono": telefono_normalizado,
                "cargo": cargo,
                "area": area,
                "estado": estado,
                "detalle": detalle,
                "email_status": email_status,
            }
        )
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=IMPORT_RESULT_COLUMNS)
    writer.writeheader()
    for result in resultados:
        writer.writerow({column: result.get(column, "") for column in IMPORT_RESULT_COLUMNS})
    return _render_admin_usuarios(
        request,
        user,
        import_summary={
            "creados": creados,
            "duplicados": duplicados,
            "errores": errores,
            "correos_enviados": correos_enviados,
            "correos_fallidos": correos_fallidos,
        },
        import_results=resultados,
        results_csv=output.getvalue(),
    )


@router.post("/admin/usuarios/importar/resultados")
def admin_descargar_import_resultados(
    request: Request,
    results_csv: str = Form(...),
    csrf_token: str = Form(...),
    user=Depends(get_current_user),
):
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not can_manage_users(user.get("rol", "")):
        return RedirectResponse("/", status_code=303)
    validate_csrf_token(request, csrf_token)
    return Response(
        content=results_csv,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="resultado-importacion-usuarios.csv"'},
    )


# ──────────────────────────────────────────────────────────────
# ANALYTICS
# ──────────────────────────────────────────────────────────────

@router.get("/admin/analytics")
def admin_analytics(request: Request, user=Depends(get_current_user)):
    if not user or user.get("rol") not in ("admin", "supervisor"):
        return RedirectResponse("/login", status_code=303)

    metrics = {
        "total_usuarios": 0,
        "usuarios_activos": 0,
        "total_art": 0,
        "art_aprobadas": 0,
        "art_rechazadas": 0,
        "art_pendientes": 0,
        "tickets_abiertos": 0,
        "tickets_en_progreso": 0,
        "tickets_resueltos": 0,
        "login_bloqueados": 0,
        "login_fallidos": 0,
        "art_por_mes": [],
        "eventos_recientes": [],
    }

    try:
        conn = _connect()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM users")
                metrics["total_usuarios"] = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM users WHERE estado_cuenta = 'activo'")
                metrics["usuarios_activos"] = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM art_records")
                metrics["total_art"] = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM art_records WHERE estado = 'aprobada'")
                metrics["art_aprobadas"] = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM art_records WHERE estado = 'rechazada'")
                metrics["art_rechazadas"] = cur.fetchone()[0]
                cur.execute("SELECT COUNT(*) FROM art_records WHERE estado = 'pendiente'")
                metrics["art_pendientes"] = cur.fetchone()[0]
                try:
                    cur.execute("SELECT COUNT(*) FROM support_tickets WHERE status = 'open'")
                    metrics["tickets_abiertos"] = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM support_tickets WHERE status = 'in_progress'")
                    metrics["tickets_en_progreso"] = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM support_tickets WHERE status = 'resolved'")
                    metrics["tickets_resueltos"] = cur.fetchone()[0]
                except Exception:
                    pass
                try:
                    cur.execute("SELECT COUNT(*) FROM system_logs WHERE event_type = 'AUTH_LOGIN_BLOCKED'")
                    metrics["login_bloqueados"] = cur.fetchone()[0]
                    cur.execute("SELECT COUNT(*) FROM system_logs WHERE event_type = 'AUTH_LOGIN_FAIL'")
                    metrics["login_fallidos"] = cur.fetchone()[0]
                    cur.execute(
                        """
                        SELECT TO_CHAR(DATE_TRUNC('month', creado_en::timestamp), 'Mon YYYY') AS mes,
                               COUNT(*) AS total
                        FROM art_records
                        WHERE creado_en::timestamp >= NOW() - INTERVAL '6 months'
                        GROUP BY DATE_TRUNC('month', creado_en::timestamp)
                        ORDER BY DATE_TRUNC('month', creado_en::timestamp)
                        """
                    )
                    metrics["art_por_mes"] = [{"mes": r[0], "total": r[1]} for r in cur.fetchall()]
                    cur.execute(
                        "SELECT event_type, username, created_at FROM system_logs"
                        " ORDER BY created_at DESC LIMIT 20"
                    )
                    metrics["eventos_recientes"] = [
                        {"event_type": r[0], "username": r[1], "created_at": r[2]}
                        for r in cur.fetchall()
                    ]
                except Exception:
                    pass
        finally:
            conn.close()
    except Exception as exc:
        logger.error("analytics_query_error: %s", exc)

    return templates.TemplateResponse(
        request,
        "admin_analytics.html",
        {"request": request, "user": user, **metrics},
    )


# ──────────────────────────────────────────────────────────────
# ERROR LOG VIEWER
# ──────────────────────────────────────────────────────────────

@router.get("/admin/errors")
def admin_errors(request: Request, user=Depends(get_current_user)):
    if not user or user.get("rol") != "admin":
        return RedirectResponse("/login", status_code=303)

    error_list = []
    try:
        from backend.src.services.logging_service import get_error_logs
        error_list = get_error_logs(limit=100)
    except Exception as exc:
        logger.error("fetch_error_logs_failed: %s", exc)

    return templates.TemplateResponse(
        request,
        "admin_errors.html",
        {"request": request, "user": user, "errors": error_list, "total": len(error_list)},
    )
